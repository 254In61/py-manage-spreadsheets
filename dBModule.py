"""Author: Allan Maseghe [allandavidke@gmail.com]"""
"""Modules to re-use in the main scripts"""
"""Created to remove the clutter on the main script too"""
"""REF: https://openpyxl.readthedocs.io/en/stable/tutorial.html"""

from openpyxl import load_workbook
from datetime import datetime

class CreateVars():
    """Class to create variables only"""
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def vars(self):
        w_book = load_workbook(self.file_name)
        """Once you gave a worksheet a name, you can get it as a key
           of the workbook. Example >>> ws3 = wb["New Title"]"""
        w_sheet = w_book[self.sheet_name]
        
        return w_book, w_sheet

class DbSearch():
    """Class to do search of values only"""
    def __init__(self,first_name,names_list,w_sheet):
        self.first_name = first_name
        self.names_list = names_list
        self.w_sheet = w_sheet
        self.process()
    
    def process(self):
        row_num = self.names_list.index(self.first_name) + 2
        DIGIT = int(0)
        print("\nClient first name:",self.first_name)
        print("\nALERT!!CORRECT MISTAKES USING MENU OPTION 5\n")

        for c in range(1,self.w_sheet.max_column + 1):
            DIGIT = DIGIT + 1
            print (DIGIT,":",self.w_sheet.cell(row=1,column=c).value, "=",
                   self.w_sheet.cell(row=row_num,column=c ).value)
        
        print("\n==================================================\n")
        

class ListsClass():
    """Return a list of first names"""
    def __init__(self,w_sheet):
        self.w_sheet = w_sheet

    def names(self):
        print("\n=======Current list of First Names=====\n")
        names_list = [self.w_sheet.cell(row=r,column=2).value
                 for r in range(2, self.w_sheet.max_row + 1)]
        print(names_list)
        print("=======================================\n")

        return names_list

class DuplicateInfo():
    """Duplicates column 1-4 of 'main' to 'services' """
    """Creating same w_sheet,w_book variables since if in different class
    seems not to save the changes"""
    def __init__(self,first_name,names_list,random_marker):
        self.first_name = first_name
        self.names_list = names_list
        self.random_marker = random_marker
        self.duplicate()

    def duplicate(self):
        w_book = load_workbook("walkin.xlsx")
        w_sheet1 = w_book["main"]
        w_sheet2 = w_book["services"]
        row_num = self.names_list.index(self.first_name) + 2
        
        if self.random_marker == "s_main":
            print("Duplicating changes on main sheet to services sheet")
            for column_num in range(1,5):
                w_sheet2.cell(row=row_num,column=column_num).value \
                = w_sheet1.cell(row=row_num,column=column_num).value

        elif self.random_marker == "s_services":
            print("Duplicating changes on services sheet to main sheet")
            for column_num in range(1,5):
                w_sheet1.cell(row=row_num,column=column_num).value \
                = w_sheet2.cell(row=row_num,column=column_num).value
        
        w_book.save("walkin.xlsx")
        w_book.close()
        print("SUCCESS!! first four columns duplicated on both sheets")
    

class EditInfo():
    """Update existing client details in walkin.xlsx"""
    """Creating same w_sheet,w_book variables since if in different class
    seems not to save the changes"""
    def __init__(self,first_name,names_list,sheet_name):
        self.first_name = first_name
        self.names_list = names_list
        self.sheet_name = sheet_name
        self.process()

    def process(self):
        w_book = load_workbook("walkin.xlsx")
        w_sheet = w_book[self.sheet_name]
        row_num = self.names_list.index(self.first_name) + 2
        
        headings_list = [w_sheet.cell(row=1,column=c).value
                    for c in range(1,w_sheet.max_column + 1)]

        heading = headings_list[int(input("Line Number to edit:")) - int(1)]
        
        if heading == "Date Registered [d/m/yr]" or heading == "Client Unique Number":
            print("ALERT!!",heading, "should not be changed")

        else:
            date_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            column_num = headings_list.index(heading) + 1
            if heading == "Comments(Additional information touching on the client)":
                current = w_sheet.cell(row=row_num,column=column_num).value
                info = ": UPDATES ON " + date_time + ":" + input(heading +":")
                w_sheet.cell(row=row_num,column=column_num).value = \
                current + info

            else:
                w_sheet.cell(row=row_num,column=column_num).value = \
                input(heading +":")
                
            w_book.save("walkin.xlsx")
            w_book.close()
            print("SUCCESS!! edited and saved")

class NewClient():
    """ Adds new clients"""
    """Creating same w_sheet,w_book variables since if in different class
    seems not to save the changes"""
    def __init__(self,empty_param):
        self.empty_param = empty_param
        self.details()

    def details(self):
        w_book = load_workbook("walkin.xlsx")
        w_sheet1 = w_book["main"]
        w_sheet2 = w_book["services"]
        r = w_sheet1.max_row + 1  

        print("\n************ADDING NEW CLIENT**************\n")

        """Would have loved to get column 2 and 3 as part of the loop!!
        But I need to match columns with website.xlsx.2nd column is the
        first name.This way, I can re-use the DbSearch module without
        much complexities"""
        
        w_sheet1.cell(row=r,column=1).value = \
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        w_sheet1.cell(row=r,column=2).value = \
        input("First name: ").capitalize()
        
        w_sheet1.cell(row=r,column=3).value = \
        input("Other names: ").capitalize()                                        

        w_sheet1.cell(row=r,column=4).value = \
        "GLAPDOO" + str(w_sheet1.max_row-1)
        
        for column_num in range(5,w_sheet1.max_column +1):
            #+1 since range excludes the last digit
            new_value = \
            input(str(w_sheet1.cell(row=1,column=column_num).value)
            + "[Press Enter for Blank value]:")

            if new_value == "":
                w_sheet1.cell(row=r,column=column_num).value = "No Record"

            else:
                w_sheet1.cell(row=r,column=column_num).value = new_value


        """Duplicate main sheet 1-4th column into services sheet """
        
        for column_num in range(1,5):
            w_sheet2.cell(row=r,column=column_num).value \
            = w_sheet1.cell(row=r,column=column_num).value


        print("\n>>>>>>>>>Input new client's service needs<<<<<<<<<<<\n")
        print("\nALERT!!Correct mistakes using Menu option 5\n")
        
        for column_num in range(5,w_sheet2.max_column + 1):
            new_info = \
            input(str(w_sheet2.cell(row=1,column=column_num).value)
            + "[Press Enter for Blank value]:")

            if new_info == "":
                w_sheet2.cell(row=r,column=column_num).value = "No Record"

            else:
                w_sheet2.cell(row=r,column=column_num).value = new_info

        w_book.save("walkin.xlsx")
        w_book.close()

        print("\nSUCCESS!!New client details saved in the walkin database.\
              \nConfirm details accuracy by using Menu option 1\
              \nALERT!!Correct mistakes using Menu option 5\n")
        print("\n*******************END*************************")




       
